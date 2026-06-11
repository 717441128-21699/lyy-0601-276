from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

from models import Contract, RiskLevel, IssueType, IssueStatus


def format_date(d: Optional[date]) -> str:
    return d.strftime("%Y-%m-%d") if d else ""


def payment_method_label(method) -> str:
    labels = {
        "monthly": "月付",
        "quarterly": "季付",
        "semi_annual": "半年付",
        "annual": "年付",
        "other": "其他",
    }
    return labels.get(getattr(method, "value", method), "其他")


def risk_level_label(level) -> str:
    labels = {"low": "低", "medium": "中", "high": "高"}
    return labels.get(getattr(level, "value", level), "低")


def issue_status_label(status) -> str:
    labels = {"pending": "待处理", "confirmed": "已确认", "resolved": "已补充", "ignored": "已忽略"}
    return labels.get(getattr(status, "value", status), "未知")


def issue_type_label(it) -> str:
    labels = {
        "missing_signature": "缺少签名",
        "date_conflict": "日期冲突",
        "lease_overlap": "租期重叠",
        "abnormal_rent": "租金异常",
        "deposit_mismatch": "押金不一致",
        "invalid_id_number": "身份证号问题",
    }
    return labels.get(getattr(it, "value", it), str(it))


def filter_contracts(
    contracts: List[Contract],
    room: Optional[str] = None,
    expire_month: Optional[str] = None,
    agent: Optional[str] = None,
    issue_type: Optional[IssueType] = None,
    high_risk_only: bool = False,
    batch_id: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[IssueStatus] = None,
) -> List[Contract]:
    result = contracts

    if room:
        result = [c for c in result if room.lower() in (c.room_number or "").lower()]
    if expire_month:
        try:
            year, month = map(int, expire_month.split("-"))
            result = [c for c in result if c.end_date and c.end_date.year == year and c.end_date.month == month]
        except ValueError:
            pass
    if agent:
        result = [c for c in result if agent.lower() in (c.agent_name or "").lower()]
    if issue_type:
        result = [c for c in result if any(i.issue_type == issue_type for i in c.issues)]
    if high_risk_only:
        result = [c for c in result if c.risk_level == RiskLevel.HIGH]
    if batch_id:
        result = [c for c in result if c.scan_batch_id == batch_id]
    if start_date:
        result = [c for c in result if c.start_date and c.start_date >= start_date]
    if end_date:
        result = [c for c in result if c.end_date and c.end_date <= end_date]
    if status:
        result = [c for c in result if any(i.status == status for i in c.issues)]

    return result


def sort_contracts(contracts: List[Contract], sort_by: str) -> List[Contract]:
    if sort_by == "risk":
        risk_order = {RiskLevel.HIGH: 0, RiskLevel.MEDIUM: 1, RiskLevel.LOW: 2}
        return sorted(contracts, key=lambda c: (risk_order.get(c.risk_level, 3), c.room_number or ""))
    elif sort_by == "end_date":
        return sorted(contracts, key=lambda c: (c.end_date or date.max, c.room_number or ""))
    elif sort_by == "room":
        return sorted(contracts, key=lambda c: (c.room_number or "", c.end_date or date.max))
    elif sort_by == "rent":
        return sorted(contracts, key=lambda c: (-c.monthly_rent, c.room_number or ""))
    return contracts


def get_pending_items(contracts: List[Contract]) -> List[dict]:
    items = []
    for c in contracts:
        missing_fields = []
        if not c.room_number:
            missing_fields.append("房号")
        if not c.tenant_name:
            missing_fields.append("租客姓名")
        if not c.tenant_id_number:
            missing_fields.append("租客身份证号")
        if not c.start_date:
            missing_fields.append("租期开始日期")
        if not c.end_date:
            missing_fields.append("租期结束日期")
        if not c.monthly_rent:
            missing_fields.append("月租金")
        if not c.deposit:
            missing_fields.append("押金")
        if not c.has_tenant_signature:
            missing_fields.append("承租方签名")
        if not c.has_landlord_signature:
            missing_fields.append("出租方签名")
        if not c.has_agent_signature:
            missing_fields.append("经纪人签名")

        pending_issues = [i for i in c.issues if i.status == IssueStatus.PENDING]
        high_risk_pending = [i for i in pending_issues if i.risk_level == RiskLevel.HIGH]

        if missing_fields or high_risk_pending:
            items.append({
                "合同文件": Path(c.file_path).name,
                "房号": c.room_number or "待补充",
                "租客": c.tenant_name or "待补充",
                "经纪人": c.agent_name or "未指定",
                "缺失字段": "、".join(missing_fields) if missing_fields else "无",
                "待处理问题数": len(pending_issues),
                "高风险问题数": len(high_risk_pending),
                "问题摘要": "; ".join(i.description for i in high_risk_pending) or "无",
            })
    return items


def get_expiring_contracts(contracts: List[Contract], days: int = 30) -> List[dict]:
    today = date.today()
    cutoff = today + timedelta(days=days)
    items = []
    for c in contracts:
        if c.end_date and today <= c.end_date <= cutoff:
            days_left = (c.end_date - today).days
            items.append({
                "合同文件": Path(c.file_path).name,
                "房号": c.room_number,
                "租客": c.tenant_name,
                "经纪人": c.agent_name or "未指定",
                "到期日期": format_date(c.end_date),
                "剩余天数": days_left,
                "月租金(元)": c.monthly_rent,
                "押金(元)": c.deposit,
                "风险等级": risk_level_label(c.risk_level),
            })
    items.sort(key=lambda x: x["剩余天数"])
    return items


def get_contract_summary(contracts: List[Contract]) -> List[dict]:
    items = []
    for c in contracts:
        items.append({
            "合同文件": Path(c.file_path).name,
            "房号": c.room_number,
            "租客": c.tenant_name,
            "出租方": c.landlord_name or "未填写",
            "经纪人": c.agent_name or "未指定",
            "租期开始": format_date(c.start_date),
            "租期结束": format_date(c.end_date),
            "月租金(元)": c.monthly_rent,
            "押金(元)": c.deposit,
            "付款方式": payment_method_label(c.payment_method),
            "风险等级": risk_level_label(c.risk_level),
            "待处理问题": c.pending_issues_count,
            "已处理问题": c.resolved_issues_count,
            "问题总数": len(c.issues),
        })
    return items


def get_review_progress(contracts: List[Contract]) -> List[dict]:
    items = []
    for c in contracts:
        if not c.issues:
            continue
        pending = c.pending_issues_count
        resolved = c.resolved_issues_count
        ignored = sum(1 for i in c.issues if i.status == IssueStatus.IGNORED)
        total = len(c.issues)
        progress = (resolved + ignored) / total * 100 if total > 0 else 100.0

        items.append({
            "合同文件": Path(c.file_path).name,
            "房号": c.room_number,
            "租客": c.tenant_name,
            "经纪人": c.agent_name or "未指定",
            "风险等级": risk_level_label(c.risk_level),
            "问题总数": total,
            "待处理": pending,
            "已确认/已补充": resolved,
            "已忽略": ignored,
            "处理进度(%)": round(progress, 1),
            "备注": c.review_notes,
        })
    items.sort(key=lambda x: x["处理进度(%)"])
    return items


def write_csv(rows: List[dict], output_path: str) -> None:
    if not rows:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("无数据\n")
        return
    fieldnames = list(rows[0].keys())
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel(rows: List[dict], output_path: str, sheet_name: str = "Sheet1") -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        write_csv(rows, str(Path(output_path).with_suffix(".csv")))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        ws["A1"] = "无数据"
        wb.save(output_path)
        return

    fieldnames = list(rows[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[name])

    for col_idx in range(1, len(fieldnames) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    wb.save(output_path)


def export_pending_list(contracts: List[Contract], output_path: str) -> str:
    items = get_pending_items(contracts)
    if output_path.endswith(".xlsx"):
        write_excel(items, output_path, "待补充清单")
    elif output_path.endswith(".csv"):
        write_csv(items, output_path)
    else:
        write_excel(items, output_path, "待补充清单")
    return output_path


def export_expiring_list(contracts: List[Contract], output_path: str, days: int = 30) -> str:
    items = get_expiring_contracts(contracts, days)
    if output_path.endswith(".xlsx"):
        write_excel(items, output_path, f"{days}天内到期清单")
    elif output_path.endswith(".csv"):
        write_csv(items, output_path)
    else:
        write_excel(items, output_path, f"{days}天内到期清单")
    return output_path


def export_contract_summary(contracts: List[Contract], output_path: str) -> str:
    items = get_contract_summary(contracts)
    if output_path.endswith(".xlsx"):
        write_excel(items, output_path, "合同摘要")
    elif output_path.endswith(".csv"):
        write_csv(items, output_path)
    else:
        write_excel(items, output_path, "合同摘要")
    return output_path


def export_review_progress(contracts: List[Contract], output_path: str) -> str:
    items = get_review_progress(contracts)
    if output_path.endswith(".xlsx"):
        write_excel(items, output_path, "处理进度表")
    elif output_path.endswith(".csv"):
        write_csv(items, output_path)
    else:
        write_excel(items, output_path, "处理进度表")
    return output_path


def export_all(contracts: List[Contract], output_dir: str, days: int = 30) -> List[str]:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outputs = []
    outputs.append(export_pending_list(contracts, str(Path(output_dir) / f"待补充清单_{timestamp}.xlsx")))
    outputs.append(export_expiring_list(contracts, str(Path(output_dir) / f"到期清单_{days}天_{timestamp}.xlsx"), days))
    outputs.append(export_contract_summary(contracts, str(Path(output_dir) / f"合同摘要_{timestamp}.xlsx")))
    outputs.append(export_review_progress(contracts, str(Path(output_dir) / f"处理进度表_{timestamp}.xlsx")))
    return outputs
